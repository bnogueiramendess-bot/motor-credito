from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
import re
from typing import Any
import unicodedata

from app.services.ar_aging_import.normalizer import as_optional_string, normalize_cnpj, normalize_money

REQUIRED_SHEETS = ("data_total", "clientes_consolidados", "ar_slide_bod")


def _normalize_label(value: str) -> str:
    cleaned = unicodedata.normalize("NFKD", value or "")
    cleaned = "".join(ch for ch in cleaned if not unicodedata.combining(ch))
    cleaned = cleaned.lower().strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def _resolve_required_sheets(sheetnames: list[str]) -> dict[str, str]:
    normalized = {_normalize_label(name): name for name in sheetnames}

    aliases = {
        "data_total": ("data total",),
        "clientes_consolidados": ("clientes consolidados", "clientes consolidado"),
        "ar_slide_bod": ("ar - slide bod",),
    }

    resolved: dict[str, str] = {}
    for key, options in aliases.items():
        for option in options:
            if option in normalized:
                resolved[key] = normalized[option]
                break
        if key in resolved:
            continue
        for normalized_name, original_name in normalized.items():
            if any(option in normalized_name for option in options):
                resolved[key] = original_name
                break
    return resolved


@dataclass(slots=True)
class ParsedAgingWorkbook:
    base_date: date
    data_total_rows: list[dict[str, Any]]
    consolidated_rows: list[dict[str, Any]]
    remark_rows: list[dict[str, Any]]
    bod_snapshot: dict[str, Any]
    bod_customer_rows: list[dict[str, Any]]
    warnings: list[str]


def extract_base_date_from_filename(filename: str) -> date:
    # Preferred format: DDMMAAAA, but also support DDMMAA (assumed 2000+YY).
    match_8 = re.search(r"(?<!\d)(\d{2})(\d{2})(\d{4})(?!\d)", filename)
    if match_8:
        day, month, year = match_8.groups()
        return date(int(year), int(month), int(day))

    match_6 = re.search(r"(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)", filename)
    if match_6:
        day, month, year_2d = match_6.groups()
        return date(2000 + int(year_2d), int(month), int(day))

    raise ValueError("Nao foi possivel extrair a data-base do nome do arquivo.")


def _extract_base_date_or_today(filename: str, warnings: list[str]) -> date:
    try:
        return extract_base_date_from_filename(filename)
    except ValueError:
        warnings.append("Data-base nao encontrada no nome do arquivo; utilizada data atual.")
        return date.today()


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = as_optional_string(value)
    if text is None:
        return None
    return text


def _row_to_raw_payload(row: tuple[Any, ...]) -> dict[str, Any]:
    return {f"col_{i + 1}": _normalize_cell(value) for i, value in enumerate(row) if _normalize_cell(value) is not None}


def _normalized_text(value: Any) -> str:
    raw = as_optional_string(value)
    if raw is None:
        return ""
    return _normalize_label(raw)


def _extract_numbers_from_row(row: tuple[Any, ...]) -> list[tuple[int, Any]]:
    values: list[tuple[int, Any]] = []
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        if isinstance(cell, (int, float)):
            values.append((idx, cell))
            continue
        if normalize_money(cell) is not None:
            values.append((idx, cell))
    return values


def _bucket_label(text: str) -> str | None:
    if not text:
        return None
    compact = text.replace(" ", "")
    if any(key in compact for key in ("0-30", "0a30", "0ate30", "ate30", "0–30")):
        return "0-30"
    if any(key in compact for key in ("1-30", "1a30", "1ate30", "1–30")):
        return "1-30"
    if any(key in compact for key in ("31-60", "31a60", "31–60")):
        return "31-60"
    if any(key in compact for key in ("61-90", "61a90", "61–90")):
        return "61-90"
    if "90+" in compact or ">90" in compact or "90dias+" in compact:
        return "90+"
    return None


def _display_bucket_label(key: str) -> str:
    mapping = {
        "0-30": "0–30 dias",
        "1-30": "1–30 dias",
        "31-60": "31–60 dias",
        "61-90": "61–90 dias",
        "90+": "90+ dias",
    }
    return mapping.get(key, key)


def _scan_bod_sheet(rows: list[tuple[Any, ...]]) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    raw_rows = [_row_to_raw_payload(row) for row in rows if any(cell is not None and str(cell).strip() != "" for cell in row)]

    risk: dict[str, dict[str, Any]] = {
        "probable": {"amount": None, "customers_count": None},
        "possible": {"amount": None, "customers_count": None},
        "rare": {"amount": None, "customers_count": None},
    }
    totals: dict[str, Any] = {}
    not_due_buckets: list[dict[str, Any]] = []
    overdue_buckets: list[dict[str, Any]] = []
    customer_rows: list[dict[str, Any]] = []
    active_bucket_section: str | None = None

    for row_idx, row in enumerate(rows, start=1):
        normalized_cells = [_normalized_text(cell) for cell in row]
        joined = " | ".join(item for item in normalized_cells if item)
        number_cells = _extract_numbers_from_row(row)

        if "not due" in joined or "a vencer" in joined:
            active_bucket_section = "not_due"
        if "overdue" in joined or "vencido" in joined:
            active_bucket_section = "overdue"

        for category in ("probable", "possible", "rare"):
            if category in joined:
                amount = None
                customers_count = None
                for _, raw in number_cells:
                    money = normalize_money(raw)
                    if money is not None and amount is None:
                        amount = money
                        continue
                    maybe_int = normalize_money(raw)
                    if maybe_int is not None and customers_count is None:
                        as_int = int(maybe_int)
                        if as_int >= 0 and as_int <= 1_000_000:
                            customers_count = as_int
                if amount is not None:
                    risk[category]["amount"] = amount
                if customers_count is not None:
                    risk[category]["customers_count"] = customers_count

        for token in normalized_cells:
            label = _bucket_label(token)
            if label is None:
                continue
            amount = None
            for _, raw in number_cells:
                parsed = normalize_money(raw)
                if parsed is not None:
                    amount = parsed
                    break
            if amount is None:
                continue
            bucket_payload = {"label": label, "amount": amount}
            if active_bucket_section == "not_due":
                not_due_buckets.append(bucket_payload)
            elif active_bucket_section == "overdue":
                overdue_buckets.append(bucket_payload)

        if "total" in joined or "open" in joined or "aging" in joined or "insured" in joined or "exposure" in joined:
            numeric_values = [normalize_money(raw) for _, raw in number_cells]
            numeric_values = [value for value in numeric_values if value is not None]
            if numeric_values:
                totals[f"row_{row_idx}"] = {
                    "text": joined[:200],
                    "values": numeric_values,
                }

        has_risk_category = any(category in joined for category in ("probable", "possible", "rare"))
        if has_risk_category:
            customer_name = as_optional_string(row[1] if len(row) > 1 else None) or as_optional_string(row[0] if len(row) > 0 else None)
            if customer_name and _normalize_label(customer_name) != "cliente":
                cnpj_candidate = next((normalize_cnpj(cell) for cell in row if normalize_cnpj(cell) is not None), None)
                money_values = [normalize_money(cell) for cell in row]
                money_values = [value for value in money_values if value is not None]
                customer_rows.append(
                    {
                        "row_number": row_idx,
                        "customer_name": customer_name,
                        "customer_document": cnpj_candidate,
                        "group_name": as_optional_string(row[2] if len(row) > 2 else None),
                        "risk_category": "probable" if "probable" in joined else "possible" if "possible" in joined else "rare",
                        "total_open_amount": money_values[0] if len(money_values) > 0 else None,
                        "overdue_amount": money_values[1] if len(money_values) > 1 else None,
                        "not_due_amount": money_values[2] if len(money_values) > 2 else None,
                        "insured_limit_amount": money_values[3] if len(money_values) > 3 else None,
                        "exposure_amount": money_values[4] if len(money_values) > 4 else None,
                        "aging_json": {},
                        "remarks": [],
                        "raw_row": _row_to_raw_payload(row),
                    }
                )

    def _dedupe_buckets(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        ordered_labels = ["0-30", "1-30", "31-60", "61-90", "90+"]
        mapped: dict[str, Any] = {}
        for item in items:
            mapped[item["label"]] = item["amount"]
        return [{"label": _display_bucket_label(label), "amount": mapped[label]} for label in ordered_labels if label in mapped]

    snapshot = {
        "risk": risk,
        "aging_buckets": {
            "not_due": _dedupe_buckets(not_due_buckets),
            "overdue": _dedupe_buckets(overdue_buckets),
        },
        "totals": totals,
        "raw_bod_json": {"rows": raw_rows},
        "warnings": warnings,
    }
    return snapshot, customer_rows, raw_rows, warnings


def _parse_aging_days(value: Any) -> int | None:
    text = as_optional_string(value)
    if text is None:
        return None
    digits = re.findall(r"\d+", text)
    if not digits:
        return None
    if len(digits) == 1:
        return int(digits[0])
    # Labels like 31-60 / 61-90: take upper bound as conservative aging reference.
    return int(digits[-1])


def _bucket_by_days(days: int, due_kind: str) -> str:
    if days <= 30:
        return "0-30" if due_kind == "not_due" else "1-30"
    if days <= 60:
        return "31-60"
    if days <= 90:
        return "61-90"
    return "90+"


def _derive_buckets_from_data_total(data_total_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    not_due_labels = ["0-30", "31-60", "61-90", "90+"]
    overdue_labels = ["1-30", "31-60", "61-90", "90+"]
    not_due_totals = {label: 0 for label in not_due_labels}
    overdue_totals = {label: 0 for label in overdue_labels}

    for row in data_total_rows:
        days = _parse_aging_days(row.get("aging"))
        if days is None:
            continue

        due_value = normalize_money(row.get("due_amount"))
        overdue_value = normalize_money(row.get("overdue_amount"))

        if due_value is not None and due_value > 0:
            label = _bucket_by_days(days, "not_due")
            not_due_totals[label] += due_value

        if overdue_value is not None and overdue_value > 0:
            label = _bucket_by_days(days, "overdue")
            overdue_totals[label] += overdue_value

    not_due = [{"label": _display_bucket_label(label), "amount": amount} for label, amount in not_due_totals.items() if amount > 0]
    overdue = [{"label": _display_bucket_label(label), "amount": amount} for label, amount in overdue_totals.items() if amount > 0]
    return not_due, overdue


def _derive_buckets_from_consolidated_sheet(consolidated_raw: list[tuple[Any, ...]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Extracts buckets from 'Clientes Consolidado' using structural columns:
    - Overdue buckets: I:O (1-30, 31-60, 61-90, 91-120, 121-180, 181-360, Above 360)
    - Not due buckets: R:X (same ranges)
    """
    overdue_sources = [
        ("1–30 dias", [8]),
        ("31–60 dias", [9]),
        ("61–90 dias", [10]),
        ("90+ dias", [11, 12, 13, 14]),
    ]
    not_due_sources = [
        ("0–30 dias", [17]),
        ("31–60 dias", [18]),
        ("61–90 dias", [19]),
        ("90+ dias", [20, 21, 22, 23]),
    ]

    best_row: tuple[Any, ...] | None = None
    best_score = -1
    for row in consolidated_raw[:8]:
        score = 0
        for idx in [8, 9, 10, 11, 12, 13, 14, 17, 18, 19, 20, 21, 22, 23]:
            if idx < len(row) and normalize_money(row[idx]) is not None:
                score += 1
        if score > best_score:
            best_score = score
            best_row = row

    if best_row is None or best_score < 6:
        return [], []

    def _sum_indices(row: tuple[Any, ...], indices: list[int]) -> Any:
        total = 0
        for idx in indices:
            if idx >= len(row):
                continue
            parsed = normalize_money(row[idx])
            if parsed is not None:
                total += parsed
        return total

    overdue: list[dict[str, Any]] = []
    not_due: list[dict[str, Any]] = []

    for label, indices in overdue_sources:
        amount = _sum_indices(best_row, indices)
        if amount > 0:
            overdue.append({"label": label, "amount": amount})

    for label, indices in not_due_sources:
        amount = _sum_indices(best_row, indices)
        if amount > 0:
            not_due.append({"label": label, "amount": amount})

    return not_due, overdue


def _find_header_row(rows: list[tuple[Any, ...]], aliases: dict[str, tuple[str, ...]]) -> tuple[int, dict[str, int]]:
    best_idx = 0
    best_mapping: dict[str, int] = {}
    best_score = -1

    for idx, row in enumerate(rows[:30]):
        lower = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        mapping: dict[str, int] = {}
        for key, options in aliases.items():
            for col_index, value in enumerate(lower):
                if any(option in value for option in options):
                    mapping[key] = col_index
                    break

        score = len(mapping)
        if score > best_score:
            best_idx = idx
            best_mapping = mapping
            best_score = score

    return best_idx, best_mapping


def _cell(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _pick_with_fallback(row: tuple[Any, ...], mapping: dict[str, int], key: str, fallback_index: int) -> Any:
    mapped_value = _cell(row, mapping.get(key))
    if mapped_value is not None and str(mapped_value).strip() != "":
        return mapped_value
    return _cell(row, fallback_index)


def _pick_optional(row: tuple[Any, ...], mapping: dict[str, int], key: str) -> Any:
    mapped_value = _cell(row, mapping.get(key))
    if mapped_value is not None and str(mapped_value).strip() != "":
        return mapped_value
    return None


def parse_aging_workbook(file_bytes: bytes, filename: str) -> ParsedAgingWorkbook:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Dependencia openpyxl nao instalada.") from exc

    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True, keep_links=False)
    warnings: list[str] = []

    resolved_sheets = _resolve_required_sheets(list(wb.sheetnames))
    missing_sheets = [name for name in REQUIRED_SHEETS if name not in resolved_sheets]
    if missing_sheets:
        raise ValueError(f"Abas obrigatorias ausentes: {', '.join(missing_sheets)}")

    base_date = _extract_base_date_or_today(filename, warnings)

    data_total_sheet = wb[resolved_sheets["data_total"]]
    data_total_raw = list(data_total_sheet.iter_rows(values_only=True))
    dt_header_index, dt_header = _find_header_row(
        data_total_raw,
        {
            "customer_name": ("cliente", "customer"),
            "document_number": ("nf", "nf ", "nota fiscal", "documento", "doc", "invoice"),
            "due_date": ("vencimento", "due date", "dt venc", "data venc"),
            "group": ("grupo", "grupo economico", "grupo economico"),
            "open_amount": ("valor", "em aberto", "open", "saldo"),
            "due_amount": ("not due", "a vencer", "vencer"),
            "overdue_amount": ("overdue", "vencido"),
            "aging": ("aging",),
        },
    )

    data_total_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(data_total_raw, start=1):
        if idx <= dt_header_index + 1:
            continue
        cnpj = row[2] if len(row) > 2 else None
        bu = row[8] if len(row) > 8 else None
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        payload = {
            "row_number": idx,
            "cnpj": cnpj,
            "customer_name": _pick_with_fallback(row, dt_header, "customer_name", 1),
            # Data Total: coluna J = Numero NF, coluna F = Data de vencimento
            "document_number": _pick_optional(row, dt_header, "document_number") or _cell(row, 9),
            "due_date": _pick_optional(row, dt_header, "due_date") or _cell(row, 5),
            "bu": bu,
            # Data Total: coluna P = Grupo Economico (fonte oficial para vinculacao com Clientes Consolidados coluna C).
            "group": _cell(row, 15) if _cell(row, 15) is not None and str(_cell(row, 15)).strip() != "" else _pick_with_fallback(row, dt_header, "group", 15),
            # Data Total: coluna D = Valor (fonte oficial de valor em aberto).
            # Mantemos fallback apenas para planilhas legadas com coluna D vazia.
            "open_amount": _cell(row, 3) if _cell(row, 3) is not None and str(_cell(row, 3)).strip() != "" else _pick_with_fallback(row, dt_header, "open_amount", 3),
            "due_amount": _pick_with_fallback(row, dt_header, "due_amount", 11),
            "overdue_amount": _pick_with_fallback(row, dt_header, "overdue_amount", 12),
            "aging": _pick_with_fallback(row, dt_header, "aging", 13),
            "raw": _row_to_raw_payload(row),
        }
        if payload["cnpj"] is None and payload["customer_name"] is None and payload["group"] is None:
            continue
        data_total_rows.append(payload)

    consolidated_sheet = wb[resolved_sheets["clientes_consolidados"]]
    consolidated_raw = list(consolidated_sheet.iter_rows(values_only=True))
    cc_header_index, cc_header = _find_header_row(
        consolidated_raw,
        {
            "group": ("grupo", "grupo economico", "grupo economico", "customer", "cliente"),
            "overdue": ("overdue", "vencido"),
            "not_due": ("not due", "a vencer"),
            "aging": ("aging", "total ar"),
            "insured_limit": ("credit insurance", "limite segurado", "insured", "coface"),
            "approved_credit": ("approved credit", "limite aprovado"),
            "exposure": ("exposi", "credit balance"),
        },
    )

    consolidated_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(consolidated_raw, start=1):
        if idx <= cc_header_index + 1:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        payload = {
            "row_number": idx,
            "bu": row[1] if len(row) > 1 else None,
            # Clientes Consolidados: coluna C = Grupo Economico (chave de comparacao com Data Total coluna P).
            "group": _cell(row, 2) if _cell(row, 2) is not None and str(_cell(row, 2)).strip() != "" else _pick_with_fallback(row, cc_header, "group", 2),
            "total_ar": row[7] if len(row) > 7 else _pick_with_fallback(row, cc_header, "aging", 7),
            "overdue_bucket_1_30": row[8] if len(row) > 8 else None,
            "overdue_bucket_31_60": row[9] if len(row) > 9 else None,
            "overdue_bucket_61_90": row[10] if len(row) > 10 else None,
            "overdue_bucket_91_120": row[11] if len(row) > 11 else None,
            "overdue_bucket_121_180": row[12] if len(row) > 12 else None,
            "overdue_bucket_181_360": row[13] if len(row) > 13 else None,
            "overdue_bucket_above_360": row[14] if len(row) > 14 else None,
            "overdue": row[15] if len(row) > 15 else _pick_with_fallback(row, cc_header, "overdue", 15),
            "not_due": row[16] if len(row) > 16 else _pick_with_fallback(row, cc_header, "not_due", 16),
            "not_due_bucket_1_30": row[17] if len(row) > 17 else None,
            "not_due_bucket_31_60": row[18] if len(row) > 18 else None,
            "not_due_bucket_61_90": row[19] if len(row) > 19 else None,
            "not_due_bucket_91_120": row[20] if len(row) > 20 else None,
            "not_due_bucket_121_180": row[21] if len(row) > 21 else None,
            "not_due_bucket_181_360": row[22] if len(row) > 22 else None,
            "not_due_bucket_above_360": row[23] if len(row) > 23 else None,
            "aging": row[7] if len(row) > 7 else _pick_with_fallback(row, cc_header, "aging", 7),
            # Clientes Consolidados: coluna AB (indice 27) = Limite Segurado (fonte obrigatoria)
            "insured_limit": _cell(row, 27),
            # Clientes Consolidados: coluna F = Limite Total Aprovado (fonte oficial).
            "approved_credit": _cell(row, 5) if _cell(row, 5) is not None and str(_cell(row, 5)).strip() != "" else _pick_with_fallback(row, cc_header, "approved_credit", 5),
            # Clientes Consolidados: coluna AH = Exposicao (fonte prioritaria)
            "exposure": row[33] if len(row) > 33 else _pick_with_fallback(row, cc_header, "exposure", 33),
            "raw": _row_to_raw_payload(row),
        }
        if payload["group"] is None and payload["overdue"] is None and payload["insured_limit"] is None:
            continue
        consolidated_rows.append(payload)

    bod_sheet = wb[resolved_sheets["ar_slide_bod"]]
    bod_raw = list(bod_sheet.iter_rows(values_only=True))
    bod_snapshot, bod_customer_rows, _, bod_warnings = _scan_bod_sheet(bod_raw)
    warnings.extend(bod_warnings)
    not_due_buckets = bod_snapshot.get("aging_buckets", {}).get("not_due", [])
    overdue_buckets = bod_snapshot.get("aging_buckets", {}).get("overdue", [])
    if not not_due_buckets and not overdue_buckets:
        derived_not_due, derived_overdue = _derive_buckets_from_consolidated_sheet(consolidated_raw)
        if derived_not_due or derived_overdue:
            bod_snapshot["aging_buckets"] = {
                "not_due": derived_not_due,
                "overdue": derived_overdue,
            }
            derived_warning = "Buckets de aging nao encontrados na aba AR - slide BoD; derivados da aba Clientes Consolidado."
            bod_snapshot.setdefault("warnings", []).append(derived_warning)
            warnings.append(derived_warning)
        else:
            derived_not_due, derived_overdue = _derive_buckets_from_data_total(data_total_rows)
            if derived_not_due or derived_overdue:
                bod_snapshot["aging_buckets"] = {
                    "not_due": derived_not_due,
                    "overdue": derived_overdue,
                }
                derived_warning = "Buckets de aging nao encontrados na aba AR - slide BoD; derivados da aba Data Total."
                bod_snapshot.setdefault("warnings", []).append(derived_warning)
                warnings.append(derived_warning)

    remark_rows: list[dict[str, Any]] = []
    for idx, row in enumerate(bod_raw, start=1):
        remark = row[15] if len(row) > 15 else None
        if as_optional_string(remark) is None:
            continue
        if _normalize_label(str(remark)) == "remark":
            continue
        identity = row[1] if len(row) > 1 else row[0] if len(row) > 0 else None
        remark_rows.append(
            {
                "row_number": idx,
                "customer_or_group": identity,
                "remark": remark,
                "raw": _row_to_raw_payload(row),
            }
        )

    if not dt_header:
        warnings.append("Cabecalhos da aba Data Total nao foram identificados; usado fallback por coluna fixa.")
    if not cc_header:
        warnings.append("Cabecalhos da aba Clientes Consolidados nao foram identificados completamente.")

    return ParsedAgingWorkbook(
        base_date=base_date,
        data_total_rows=data_total_rows,
        consolidated_rows=consolidated_rows,
        remark_rows=remark_rows,
        bod_snapshot=bod_snapshot,
        bod_customer_rows=bod_customer_rows,
        warnings=warnings,
    )
