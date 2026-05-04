from __future__ import annotations

from pathlib import Path
from dataclasses import dataclass
from typing import Any

from app.services.ar_aging_import.normalizer import as_optional_string

SHEET_NAME = "AR - slide BoD"
START_ROW = 26
STOP_LABEL = "TOTAL OVER DUE IN LITIGATION (TOP 12)"


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = as_optional_string(value)
    if text is None or text == "-":
        return 0.0

    sanitized = text.replace("R$", "").replace(" ", "")
    if "," in sanitized and "." in sanitized:
        sanitized = sanitized.replace(".", "").replace(",", ".")
    elif "," in sanitized:
        sanitized = sanitized.replace(".", "").replace(",", ".")
    elif "." in sanitized:
        sanitized = sanitized.replace(".", "")

    try:
        return float(sanitized)
    except ValueError:
        return 0.0


def _safe_div(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return float(numerator / denominator)


def _is_valid_customer_name(value: Any) -> bool:
    name = as_optional_string(value)
    if name is None:
        return False

    normalized = " ".join(name.strip().upper().split())
    if not normalized:
        return False
    if "TOTAL" in normalized:
        return False
    if normalized == "LOW RISK - EXPECT TO RECEIVE SOON":
        return False
    return True


@dataclass(slots=True)
class RiskRow:
    customer_name: str
    bu: str | None
    remark: str | None
    exposure_amount: float
    critical_amount: float
    attention_amount: float
    healthy_amount: float
    critical_countable: bool
    attention_countable: bool
    healthy_countable: bool


def _is_countable_positive(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return float(value) > 0
    text = as_optional_string(value)
    if text is None:
        return False
    if text.strip() == "-":
        return False
    return _to_float(text) > 0


def _build_summary(rows: list[RiskRow]) -> dict[str, Any]:
    critical_amount = 0.0
    attention_amount = 0.0
    healthy_amount = 0.0

    critical_clients = 0
    attention_clients = 0
    healthy_clients = 0
    at_risk_clients = 0

    for row in rows:
        critical = row.critical_amount
        attention = row.attention_amount
        healthy = row.healthy_amount
        if critical == 0 and attention == 0 and healthy == 0:
            continue
        critical_amount += critical
        attention_amount += attention
        healthy_amount += healthy
        if row.critical_countable:
            critical_clients += 1
        if row.attention_countable:
            attention_clients += 1
        if row.healthy_countable:
            healthy_clients += 1
        if row.critical_countable or row.attention_countable:
            at_risk_clients += 1

    total_portfolio = critical_amount + attention_amount + healthy_amount
    at_risk_amount = critical_amount + attention_amount

    top_clients_candidates: list[dict[str, Any]] = []
    for row in rows:
        if row.critical_amount <= 0 and row.attention_amount <= 0:
            continue
        if row.exposure_amount <= 500_000:
            continue
        risk_level = "critical" if row.critical_amount > 0 else "attention"
        top_clients_candidates.append(
            {
                "customer_name": row.customer_name,
                "bu": row.bu,
                "remark": row.remark,
                "amount": float(row.exposure_amount),
                "risk_level": risk_level,
            }
        )
    top_clients_at_risk = sorted(
        top_clients_candidates,
        key=lambda item: -item["amount"],
    )

    return {
        "at_risk_amount": float(at_risk_amount),
        "at_risk_percentage": _safe_div(at_risk_amount, total_portfolio),
        "healthy_percentage": _safe_div(healthy_amount, total_portfolio),
        "clients_at_risk": int(at_risk_clients),
        "distribution": {
            "critical": {
                "amount": float(critical_amount),
                "percentage": _safe_div(critical_amount, total_portfolio),
                "clients": int(critical_clients),
            },
            "attention": {
                "amount": float(attention_amount),
                "percentage": _safe_div(attention_amount, total_portfolio),
                "clients": int(attention_clients),
            },
            "healthy": {
                "amount": float(healthy_amount),
                "percentage": _safe_div(healthy_amount, total_portfolio),
                "clients": int(healthy_clients),
            },
        },
        "top_clients_at_risk": top_clients_at_risk,
    }


def _parse_bod_rows(rows: list[tuple[Any, ...]]) -> list[RiskRow]:
    parsed_rows: list[RiskRow] = []
    for row_number, row in enumerate(rows, start=1):
        if row_number < START_ROW:
            continue
        customer_name = as_optional_string(row[1] if len(row) > 1 else None)
        if customer_name and customer_name.strip().upper() == STOP_LABEL:
            break
        if not _is_valid_customer_name(customer_name):
            continue

        raw_critical = row[11] if len(row) > 11 else None
        raw_attention = row[12] if len(row) > 12 else None
        raw_healthy = row[13] if len(row) > 13 else None

        parsed_rows.append(
            RiskRow(
                customer_name=customer_name,
                bu=as_optional_string(row[3] if len(row) > 3 else None),
                remark=as_optional_string(row[15] if len(row) > 15 else None),
                exposure_amount=_to_float(row[6] if len(row) > 6 else None),
                critical_amount=_to_float(raw_critical),
                attention_amount=_to_float(raw_attention),
                healthy_amount=_to_float(raw_healthy),
                critical_countable=_is_countable_positive(raw_critical),
                attention_countable=_is_countable_positive(raw_attention),
                healthy_countable=_is_countable_positive(raw_healthy),
            )
        )
    return parsed_rows


def calculate_portfolio_risk_from_bod(file_path: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Dependencia openpyxl nao instalada.") from exc

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo nao encontrado: {file_path}")

    wb = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Aba obrigatoria ausente: {SHEET_NAME}")

    sheet = wb[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    parsed_rows = _parse_bod_rows(rows)
    return _build_summary(parsed_rows)


def calculate_portfolio_risk_from_bod_raw_rows(raw_rows: list[dict[str, Any]]) -> dict[str, Any]:
    normalized_rows: list[RiskRow] = []
    for raw in raw_rows:
        if not isinstance(raw, dict):
            continue
        customer_name = as_optional_string(raw.get("col_2"))
        if customer_name and customer_name.strip().upper() == STOP_LABEL:
            break
        if not _is_valid_customer_name(customer_name):
            continue
        raw_critical = raw.get("col_12")
        raw_attention = raw.get("col_13")
        raw_healthy = raw.get("col_14")
        normalized_rows.append(
            RiskRow(
                customer_name=customer_name,
                bu=as_optional_string(raw.get("col_4")),
                remark=as_optional_string(raw.get("col_16")),
                exposure_amount=_to_float(raw.get("col_7")),
                critical_amount=_to_float(raw_critical),
                attention_amount=_to_float(raw_attention),
                healthy_amount=_to_float(raw_healthy),
                critical_countable=_is_countable_positive(raw_critical),
                attention_countable=_is_countable_positive(raw_attention),
                healthy_countable=_is_countable_positive(raw_healthy),
            )
        )
    return _build_summary(normalized_rows)
