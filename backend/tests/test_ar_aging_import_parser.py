from __future__ import annotations

from io import BytesIO
import unittest

from app.services.ar_aging_import.parser import parse_aging_workbook


class ArAgingImportParserTestCase(unittest.TestCase):
    def setUp(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl não disponível")

        wb = Workbook()

        ws_dt = wb.active
        ws_dt.title = "Data Total"
        ws_dt.append(["", "Cliente", "CNPJ", "", "", "", "", "", "BU", "Grupo", "Em Aberto", "Not Due", "Overdue", "Aging"])
        ws_dt.append(["", "Cliente A", "21.839.049/0001-02", "", "", "", "", "", "Aditivos", "Grupo A", "1.200,00", "500,00", "700,00", "31-60"])

        ws_cc = wb.create_sheet("Clientes Consolidados")
        ws_cc.append(["Grupo", "Overdue", "Not Due", "Aging", "Limite Segurado", "Exposição"])
        ws_cc.append(["Grupo A", "700,00", "500,00", "1200,00", "5000,00", "6200,00"])

        ws_bod = wb.create_sheet("AR - slide BoD")
        ws_bod.append(["A", "Cliente", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "Remark"])
        ws_bod.append(["", "Probable", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "Possible", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "Rare", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "Not Due", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "0-30 dias", "1000", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "31-60 dias", "500", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "Overdue", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "1-30 dias", "700", "", "", "", "", "", "", "", "", "", "", "", "", ""])
        ws_bod.append(["", "Grupo A", "", "", "", "", "", "", "", "", "", "", "", "", "", "Cliente com risco moderado"])

        stream = BytesIO()
        wb.save(stream)
        self.bytes_data = stream.getvalue()

    def test_parse_required_sheets(self) -> None:
        parsed = parse_aging_workbook(self.bytes_data, "27042025- AR Additive-Fertilizer_closing.xlsx")

        self.assertEqual(str(parsed.base_date), "2025-04-27")
        self.assertEqual(len(parsed.data_total_rows), 1)
        self.assertEqual(len(parsed.consolidated_rows), 1)
        self.assertEqual(len(parsed.remark_rows), 1)
        self.assertIn("risk", parsed.bod_snapshot)
        self.assertIn("aging_buckets", parsed.bod_snapshot)
        self.assertIn("raw_bod_json", parsed.bod_snapshot)

    def test_parser_does_not_depend_on_exact_filename(self) -> None:
        parsed = parse_aging_workbook(self.bytes_data, "30042026- AR Additive-Fertilizer_closing.xlsx")
        self.assertEqual(str(parsed.base_date), "2026-04-30")
        self.assertEqual(len(parsed.data_total_rows), 1)

    def test_parser_accepts_filename_without_date_using_sheet_structure(self) -> None:
        parsed = parse_aging_workbook(self.bytes_data, "aging_import.xlsx")
        self.assertGreaterEqual(len(parsed.warnings), 1)
        self.assertEqual(len(parsed.data_total_rows), 1)

    def test_fallback_derives_buckets_from_data_total_when_bod_has_no_buckets(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl não disponível")

        wb = Workbook()
        ws_dt = wb.active
        ws_dt.title = "Data Total"
        ws_dt.append(["", "Cliente", "CNPJ", "", "", "", "", "", "BU", "Grupo", "Em Aberto", "Not Due", "Overdue", "Aging"])
        ws_dt.append(["", "Cliente A", "21.839.049/0001-02", "", "", "", "", "", "Aditivos", "Grupo A", "1.200,00", "500,00", "700,00", "31"])
        ws_dt.append(["", "Cliente B", "09.326.402/0001-98", "", "", "", "", "", "Aditivos", "Grupo B", "900,00", "900,00", "0,00", "75"])

        ws_cc = wb.create_sheet("Clientes Consolidados")
        ws_cc.append(["Grupo", "Overdue", "Not Due", "Aging", "Limite Segurado", "Exposição"])
        ws_cc.append(["Grupo A", "700,00", "1400,00", "2100,00", "5000,00", "0,00"])

        ws_bod = wb.create_sheet("AR - slide BoD")
        ws_bod.append(["A", "Cliente", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "Remark"])
        ws_bod.append(["", "Resumo sem buckets explícitos", "", "", "", "", "", "", "", "", "", "", "", "", "", "Observação"])

        stream = BytesIO()
        wb.save(stream)
        bytes_data = stream.getvalue()

        parsed = parse_aging_workbook(bytes_data, "30042026- qualquer_nome.xlsx")
        not_due = parsed.bod_snapshot.get("aging_buckets", {}).get("not_due", [])
        overdue = parsed.bod_snapshot.get("aging_buckets", {}).get("overdue", [])
        warnings = parsed.bod_snapshot.get("warnings", [])

        self.assertTrue(len(not_due) > 0)
        self.assertTrue(len(overdue) > 0)
        self.assertTrue(any("derivados da aba Data Total" in warning for warning in warnings))


if __name__ == "__main__":
    unittest.main()
